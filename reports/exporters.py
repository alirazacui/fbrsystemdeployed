"""
========================================================
reports/exporters.py
 
Advanced Reports — Excel + PDF Export
 
Five report types:
  1. Daily Sales Summary
  2. Monthly Sales Summary
  3. Product-wise Sales
  4. Cashier-wise Sales
  5. FBR Submission Status
 
Each report:
  - Generates Excel (.xlsx) with formatting and formulas
  - Generates PDF using reportlab
  - Saves to S3 and returns URL
  - Scoped to company (or all companies for admin)
========================================================
"""
 
import io
import logging
from datetime import datetime, date, timedelta
from decimal import Decimal
 
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.utils import timezone
 
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
 
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable,
)
 
logger = logging.getLogger(__name__)
 
# ── Brand colors ──────────────────────────────────────────────────────────────
BLUE_HEX   = "1A56DB"
GRAY_HEX   = "F3F4F6"
WHITE_HEX  = "FFFFFF"
BLACK_HEX  = "111827"
GREEN_HEX  = "065F46"
RED_HEX    = "991B1B"
 
# openpyxl fills
HEADER_FILL  = PatternFill("solid", fgColor=BLUE_HEX)
ALT_FILL     = PatternFill("solid", fgColor="EFF6FF")
TOTAL_FILL   = PatternFill("solid", fgColor="DBEAFE")
SUBHEAD_FILL = PatternFill("solid", fgColor=GRAY_HEX)
 
# openpyxl fonts
HEADER_FONT  = Font(bold=True, color=WHITE_HEX, name="Arial", size=10)
TITLE_FONT   = Font(bold=True, color=BLACK_HEX, name="Arial", size=14)
SUBHEAD_FONT = Font(bold=True, color=BLACK_HEX, name="Arial", size=10)
TOTAL_FONT   = Font(bold=True, color=BLACK_HEX, name="Arial", size=10)
BODY_FONT    = Font(name="Arial", size=9)
 
THIN_BORDER  = Border(
    left   = Side(style="thin", color="E5E7EB"),
    right  = Side(style="thin", color="E5E7EB"),
    top    = Side(style="thin", color="E5E7EB"),
    bottom = Side(style="thin", color="E5E7EB"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN  = Alignment(horizontal="right",  vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
 
PKR_FORMAT   = '#,##0.00'
INT_FORMAT   = '#,##0'
 
 
# ---------------------------------------------------------------------------
# S3 save helper
# ---------------------------------------------------------------------------
 
def _save_to_s3(buffer: io.BytesIO, company, report_name: str, ext: str) -> str:
    """Saves buffer to S3 and returns public URL."""
    now  = timezone.now()
    path = (
        f"company_{company.pk}/reports/"
        f"{now.strftime('%Y/%m')}/"
        f"{report_name}_{now.strftime('%Y%m%d_%H%M%S')}.{ext}"
    )
    buffer.seek(0)
    saved = default_storage.save(path, ContentFile(buffer.read()))
    return default_storage.url(saved)
 
 
# ---------------------------------------------------------------------------
# Shared Excel helpers
# ---------------------------------------------------------------------------
 
def _write_report_header(ws, title: str, company_name: str,
                          date_from: date, date_to: date, col_count: int):
    """Writes company name, report title, and date range at top of sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"] = company_name
    ws["A1"].font      = Font(bold=True, name="Arial", size=12, color=BLUE_HEX)
    ws["A1"].alignment = CENTER_ALIGN
 
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"] = title
    ws["A2"].font      = TITLE_FONT
    ws["A2"].alignment = CENTER_ALIGN
 
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)
    ws["A3"] = f"Period: {date_from.strftime('%d %b %Y')} to {date_to.strftime('%d %b %Y')}"
    ws["A3"].font      = Font(name="Arial", size=9, color="6B7280")
    ws["A3"].alignment = CENTER_ALIGN
 
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 16
 
    return 5   # next available row
 
 
def _style_header_row(ws, row: int, col_count: int):
    """Applies blue header styling to a row."""
    for col in range(1, col_count + 1):
        cell            = ws.cell(row=row, column=col)
        cell.font       = HEADER_FONT
        cell.fill       = HEADER_FILL
        cell.alignment  = CENTER_ALIGN
        cell.border     = THIN_BORDER
    ws.row_dimensions[row].height = 18
 
 
def _style_data_row(ws, row: int, col_count: int, alternate: bool = False):
    """Applies alternating row styling."""
    for col in range(1, col_count + 1):
        cell           = ws.cell(row=row, column=col)
        cell.font      = BODY_FONT
        cell.border    = THIN_BORDER
        if alternate:
            cell.fill  = ALT_FILL
 
 
def _style_total_row(ws, row: int, col_count: int):
    """Applies total row styling."""
    for col in range(1, col_count + 1):
        cell           = ws.cell(row=row, column=col)
        cell.font      = TOTAL_FONT
        cell.fill      = TOTAL_FILL
        cell.border    = THIN_BORDER
 
 
def _autofit_columns(ws, min_width=10, max_width=40):
    """Auto-fits column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, min_width), max_width
        )
 
 
# ===========================================================================
# 1. DAILY SALES SUMMARY EXPORTER
# ===========================================================================
 
class DailySalesExporter:
    """Exports daily sales summary for a date range."""
 
    def __init__(self, company, date_from: date, date_to: date):
        self.company   = company
        self.date_from = date_from
        self.date_to   = date_to
        self.data      = self._fetch_data()
 
    def _fetch_data(self) -> list:
        from pos.models import Sale, SaleStatus, SalePayment, PaymentMethod
        from django.db.models import Sum, Count, Avg
        from django.db.models.functions import TruncDate
 
        dt_from = timezone.make_aware(datetime.combine(self.date_from, datetime.min.time()))
        dt_to   = timezone.make_aware(datetime.combine(self.date_to,   datetime.max.time()))
 
        daily = list(
            Sale.objects.filter(
                company            = self.company,
                status             = SaleStatus.COMPLETED,
                completed_at__range = (dt_from, dt_to),
            )
            .annotate(sale_date=TruncDate("completed_at"))
            .values("sale_date")
            .annotate(
                total_sales    = Count("id"),
                total_revenue  = Sum("total_amount"),
                total_subtotal = Sum("subtotal"),
                total_tax      = Sum("total_tax"),
                total_discount = Sum("total_discount"),
                avg_sale       = Avg("total_amount"),
            )
            .order_by("sale_date")
        )
        return daily
 
    def export_excel(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Sales"
        ws.freeze_panes = "A6"
 
        COLS = 8
        next_row = _write_report_header(
            ws, "Daily Sales Summary",
            self.company.business_name,
            self.date_from, self.date_to, COLS
        )
 
        # Headers
        headers = ["Date", "No. of Sales", "Subtotal (Rs.)",
                   "Discount (Rs.)", "Tax (Rs.)",
                   "Total Revenue (Rs.)", "Avg Sale (Rs.)", "Day"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=next_row, column=i, value=h)
        _style_header_row(ws, next_row, COLS)
        next_row += 1
 
        data_start = next_row
 
        # Data rows
        for i, row in enumerate(self.data):
            r = next_row
            d = row["sale_date"]
            ws.cell(r, 1, d.strftime("%d-%b-%Y"))
            ws.cell(r, 2, row["total_sales"])
            ws.cell(r, 3, float(row["total_subtotal"] or 0))
            ws.cell(r, 4, float(row["total_discount"] or 0))
            ws.cell(r, 5, float(row["total_tax"]      or 0))
            ws.cell(r, 6, float(row["total_revenue"]  or 0))
            ws.cell(r, 7, float(row["avg_sale"]       or 0))
            ws.cell(r, 8, d.strftime("%A"))
 
            # Number formats
            for col in [3, 4, 5, 6, 7]:
                ws.cell(r, col).number_format = PKR_FORMAT
            ws.cell(r, 2).number_format = INT_FORMAT
            ws.cell(r, 1).alignment = LEFT_ALIGN
            ws.cell(r, 8).alignment = CENTER_ALIGN
 
            _style_data_row(ws, r, COLS, alternate=(i % 2 == 1))
            next_row += 1
 
        data_end = next_row - 1
 
        # Totals row using Excel formulas
        r = next_row
        ws.cell(r, 1, "TOTAL")
        ws.cell(r, 1).alignment = CENTER_ALIGN
        ws.cell(r, 2, f"=SUM(B{data_start}:B{data_end})")
        ws.cell(r, 3, f"=SUM(C{data_start}:C{data_end})")
        ws.cell(r, 4, f"=SUM(D{data_start}:D{data_end})")
        ws.cell(r, 5, f"=SUM(E{data_start}:E{data_end})")
        ws.cell(r, 6, f"=SUM(F{data_start}:F{data_end})")
        ws.cell(r, 7, f"=AVERAGE(G{data_start}:G{data_end})")
        ws.cell(r, 8, "")
        for col in [3, 4, 5, 6, 7]:
            ws.cell(r, col).number_format = PKR_FORMAT
        ws.cell(r, 2).number_format = INT_FORMAT
        _style_total_row(ws, r, COLS)
 
        # Bar chart — Revenue by day
        if self.data:
            chart = BarChart()
            chart.title    = "Daily Revenue"
            chart.y_axis.title = "Revenue (Rs.)"
            chart.x_axis.title = "Date"
            chart.style    = 10
            chart.width    = 20
            chart.height   = 12
 
            data_ref  = Reference(ws, min_col=6, min_row=data_start - 1,
                                  max_row=data_end)
            cats_ref  = Reference(ws, min_col=1, min_row=data_start,
                                  max_row=data_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, f"A{next_row + 3}")
 
        _autofit_columns(ws)
        ws.sheet_view.showGridLines = False
 
        # Save
        buffer = io.BytesIO()
        wb.save(buffer)
        return _save_to_s3(buffer, self.company, "daily_sales", "xlsx")
 
    def export_pdf(self) -> str:
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=15*mm, leftMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm,
        )
        styles = getSampleStyleSheet()
        story  = []
        width  = landscape(A4)[0] - 30*mm
 
        # Title
        story.append(Paragraph(
            self.company.business_name,
            ParagraphStyle("co", fontSize=12, fontName="Helvetica-Bold",
                           textColor=colors.HexColor("#1A56DB"),
                           alignment=TA_CENTER)
        ))
        story.append(Paragraph(
            "Daily Sales Summary",
            ParagraphStyle("title", fontSize=16, fontName="Helvetica-Bold",
                           alignment=TA_CENTER, spaceAfter=4)
        ))
        story.append(Paragraph(
            f"Period: {self.date_from.strftime('%d %b %Y')} to "
            f"{self.date_to.strftime('%d %b %Y')}",
            ParagraphStyle("sub", fontSize=9, alignment=TA_CENTER,
                           textColor=colors.HexColor("#6B7280"), spaceAfter=8)
        ))
 
        # Table
        col_style = ParagraphStyle("ch", fontSize=8, fontName="Helvetica-Bold",
                                   textColor=colors.white, alignment=TA_CENTER)
        cell_r    = ParagraphStyle("cr", fontSize=8, alignment=TA_RIGHT)
        cell_c    = ParagraphStyle("cc", fontSize=8, alignment=TA_CENTER)
 
        table_data = [[
            Paragraph("Date",             col_style),
            Paragraph("Sales",            col_style),
            Paragraph("Subtotal",         col_style),
            Paragraph("Discount",         col_style),
            Paragraph("Tax",              col_style),
            Paragraph("Total Revenue",    col_style),
            Paragraph("Avg Sale",         col_style),
        ]]
 
        totals = [0, 0, 0, 0, 0, 0]
        for row in self.data:
            rev  = float(row["total_revenue"]  or 0)
            sub  = float(row["total_subtotal"] or 0)
            dis  = float(row["total_discount"] or 0)
            tax  = float(row["total_tax"]      or 0)
            avg  = float(row["avg_sale"]       or 0)
            cnt  = row["total_sales"]
 
            totals[0] += cnt
            totals[1] += sub
            totals[2] += dis
            totals[3] += tax
            totals[4] += rev
 
            table_data.append([
                Paragraph(row["sale_date"].strftime("%d-%b-%Y"), cell_c),
                Paragraph(str(cnt),           cell_c),
                Paragraph(f"Rs. {sub:,.2f}",  cell_r),
                Paragraph(f"Rs. {dis:,.2f}",  cell_r),
                Paragraph(f"Rs. {tax:,.2f}",  cell_r),
                Paragraph(f"Rs. {rev:,.2f}",  cell_r),
                Paragraph(f"Rs. {avg:,.2f}",  cell_r),
            ])
 
        # Totals row
        avg_total = totals[4] / totals[0] if totals[0] > 0 else 0
        tf = ParagraphStyle("tf", fontSize=8, fontName="Helvetica-Bold",
                            alignment=TA_RIGHT)
        tc = ParagraphStyle("tc", fontSize=8, fontName="Helvetica-Bold",
                            alignment=TA_CENTER)
        table_data.append([
            Paragraph("TOTAL", tc),
            Paragraph(str(totals[0]), tc),
            Paragraph(f"Rs. {totals[1]:,.2f}", tf),
            Paragraph(f"Rs. {totals[2]:,.2f}", tf),
            Paragraph(f"Rs. {totals[3]:,.2f}", tf),
            Paragraph(f"Rs. {totals[4]:,.2f}", tf),
            Paragraph(f"Rs. {avg_total:,.2f}", tf),
        ])
 
        col_w = width / 7
        t = Table(table_data, colWidths=[col_w] * 7, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#1A56DB")),
            ("BACKGROUND",    (0, -1),(-1, -1), colors.HexColor("#DBEAFE")),
            ("ROWBACKGROUNDS",(0, 1), (-1, -2),
             [colors.white, colors.HexColor("#EFF6FF")]),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
 
        # Footer
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            f"Generated: {timezone.now().strftime('%d-%b-%Y %I:%M %p')}",
            ParagraphStyle("ft", fontSize=7, alignment=TA_RIGHT,
                           textColor=colors.HexColor("#9CA3AF"))
        ))
 
        doc.build(story)
        return _save_to_s3(buffer, self.company, "daily_sales", "pdf")
 
 
# ===========================================================================
# 2. MONTHLY SALES SUMMARY EXPORTER
# ===========================================================================
 
class MonthlySalesExporter:
    """Exports monthly sales summary."""
 
    def __init__(self, company, year: int):
        self.company = company
        self.year    = year
        self.data    = self._fetch_data()
 
    def _fetch_data(self) -> list:
        from pos.models import Sale, SaleStatus
        from django.db.models import Sum, Count, Avg
        from django.db.models.functions import TruncMonth
 
        return list(
            Sale.objects.filter(
                company              = self.company,
                status               = SaleStatus.COMPLETED,
                completed_at__year   = self.year,
            )
            .annotate(month=TruncMonth("completed_at"))
            .values("month")
            .annotate(
                total_sales    = Count("id"),
                total_revenue  = Sum("total_amount"),
                total_tax      = Sum("total_tax"),
                total_discount = Sum("total_discount"),
                avg_sale       = Avg("total_amount"),
            )
            .order_by("month")
        )
 
    def export_excel(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Sales"
        ws.freeze_panes = "A6"
 
        COLS   = 6
        MONTHS = ["Jan","Feb","Mar","Apr","May","Jun",
                  "Jul","Aug","Sep","Oct","Nov","Dec"]
 
        next_row = _write_report_header(
            ws, f"Monthly Sales Summary — {self.year}",
            self.company.business_name,
            date(self.year, 1, 1), date(self.year, 12, 31), COLS
        )
 
        headers = ["Month", "No. of Sales", "Total Revenue (Rs.)",
                   "Total Tax (Rs.)", "Total Discount (Rs.)", "Avg Sale (Rs.)"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=next_row, column=i, value=h)
        _style_header_row(ws, next_row, COLS)
        next_row += 1
 
        data_start = next_row
        month_data = {r["month"].month: r for r in self.data}
 
        for m in range(1, 13):
            r   = next_row
            row = month_data.get(m)
            ws.cell(r, 1, MONTHS[m-1])
            ws.cell(r, 1).alignment = LEFT_ALIGN
            if row:
                ws.cell(r, 2, row["total_sales"])
                ws.cell(r, 3, float(row["total_revenue"]  or 0))
                ws.cell(r, 4, float(row["total_tax"]      or 0))
                ws.cell(r, 5, float(row["total_discount"] or 0))
                ws.cell(r, 6, float(row["avg_sale"]       or 0))
            else:
                for c in range(2, 7):
                    ws.cell(r, c, 0)
 
            for col in [3, 4, 5, 6]:
                ws.cell(r, col).number_format = PKR_FORMAT
            ws.cell(r, 2).number_format = INT_FORMAT
            _style_data_row(ws, r, COLS, alternate=(m % 2 == 0))
            next_row += 1
 
        data_end = next_row - 1
 
        # Totals
        r = next_row
        ws.cell(r, 1, "ANNUAL TOTAL")
        ws.cell(r, 2, f"=SUM(B{data_start}:B{data_end})")
        ws.cell(r, 3, f"=SUM(C{data_start}:C{data_end})")
        ws.cell(r, 4, f"=SUM(D{data_start}:D{data_end})")
        ws.cell(r, 5, f"=SUM(E{data_start}:E{data_end})")
        ws.cell(r, 6, f"=AVERAGE(F{data_start}:F{data_end})")
        for col in [3, 4, 5, 6]:
            ws.cell(r, col).number_format = PKR_FORMAT
        ws.cell(r, 2).number_format = INT_FORMAT
        _style_total_row(ws, r, COLS)
 
        _autofit_columns(ws)
        ws.sheet_view.showGridLines = False
 
        buffer = io.BytesIO()
        wb.save(buffer)
        return _save_to_s3(
            buffer, self.company, f"monthly_sales_{self.year}", "xlsx"
        )
 
    def export_pdf(self) -> str:
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=15*mm, leftMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm,
        )
        styles = getSampleStyleSheet()
        story  = []
        width  = A4[0] - 30*mm
        MONTHS = ["January","February","March","April","May","June",
                  "July","August","September","October","November","December"]
 
        story.append(Paragraph(
            self.company.business_name,
            ParagraphStyle("co", fontSize=12, fontName="Helvetica-Bold",
                           textColor=colors.HexColor("#1A56DB"),
                           alignment=TA_CENTER)
        ))
        story.append(Paragraph(
            f"Monthly Sales Summary — {self.year}",
            ParagraphStyle("title", fontSize=14, fontName="Helvetica-Bold",
                           alignment=TA_CENTER, spaceAfter=8)
        ))
 
        col_style = ParagraphStyle("ch", fontSize=8, fontName="Helvetica-Bold",
                                   textColor=colors.white, alignment=TA_CENTER)
        cell_r    = ParagraphStyle("cr", fontSize=8, alignment=TA_RIGHT)
        cell_c    = ParagraphStyle("cc", fontSize=8, alignment=TA_CENTER)
 
        table_data = [[
            Paragraph("Month",          col_style),
            Paragraph("Sales",          col_style),
            Paragraph("Revenue",        col_style),
            Paragraph("Tax",            col_style),
            Paragraph("Discount",       col_style),
            Paragraph("Avg Sale",       col_style),
        ]]
 
        month_data  = {r["month"].month: r for r in self.data}
        totals      = [0, 0, 0, 0, 0]
 
        for m in range(1, 13):
            row = month_data.get(m)
            if row:
                cnt = row["total_sales"]
                rev = float(row["total_revenue"]  or 0)
                tax = float(row["total_tax"]      or 0)
                dis = float(row["total_discount"] or 0)
                avg = float(row["avg_sale"]       or 0)
                totals[0] += cnt
                totals[1] += rev
                totals[2] += tax
                totals[3] += dis
            else:
                cnt = rev = tax = dis = avg = 0
 
            table_data.append([
                Paragraph(MONTHS[m-1],        cell_c),
                Paragraph(str(cnt),           cell_c),
                Paragraph(f"Rs. {rev:,.2f}",  cell_r),
                Paragraph(f"Rs. {tax:,.2f}",  cell_r),
                Paragraph(f"Rs. {dis:,.2f}",  cell_r),
                Paragraph(f"Rs. {avg:,.2f}",  cell_r),
            ])
 
        avg_total = totals[1] / totals[0] if totals[0] > 0 else 0
        tf = ParagraphStyle("tf", fontSize=8, fontName="Helvetica-Bold",
                            alignment=TA_RIGHT)
        tc = ParagraphStyle("tc", fontSize=8, fontName="Helvetica-Bold",
                            alignment=TA_CENTER)
        table_data.append([
            Paragraph("TOTAL", tc),
            Paragraph(str(totals[0]), tc),
            Paragraph(f"Rs. {totals[1]:,.2f}", tf),
            Paragraph(f"Rs. {totals[2]:,.2f}", tf),
            Paragraph(f"Rs. {totals[3]:,.2f}", tf),
            Paragraph(f"Rs. {avg_total:,.2f}", tf),
        ])
 
        col_w = width / 6
        t = Table(table_data, colWidths=[col_w*1.5] + [col_w]*5, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0),  (-1, 0),  colors.HexColor("#1A56DB")),
            ("BACKGROUND",    (0, -1), (-1, -1), colors.HexColor("#DBEAFE")),
            ("ROWBACKGROUNDS",(0, 1),  (-1, -2),
             [colors.white, colors.HexColor("#EFF6FF")]),
            ("GRID",          (0, 0),  (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",    (0, 0),  (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
            ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            f"Generated: {timezone.now().strftime('%d-%b-%Y %I:%M %p')}",
            ParagraphStyle("ft", fontSize=7, alignment=TA_RIGHT,
                           textColor=colors.HexColor("#9CA3AF"))
        ))
        doc.build(story)
        return _save_to_s3(
            buffer, self.company, f"monthly_sales_{self.year}", "pdf"
        )
 
 
# ===========================================================================
# 3. PRODUCT-WISE SALES EXPORTER
# ===========================================================================
 
class ProductSalesExporter:
 
    def __init__(self, company, date_from: date, date_to: date,
                 limit: int = 100):
        self.company   = company
        self.date_from = date_from
        self.date_to   = date_to
        self.limit     = limit
        self.data      = self._fetch_data()
 
    def _fetch_data(self) -> list:
        from pos.models import SaleLine, SaleStatus
        from django.db.models import Sum, Count
 
        dt_from = timezone.make_aware(
            datetime.combine(self.date_from, datetime.min.time())
        )
        dt_to   = timezone.make_aware(
            datetime.combine(self.date_to, datetime.max.time())
        )
 
        return list(
            SaleLine.objects.filter(
                sale__company            = self.company,
                sale__status             = SaleStatus.COMPLETED,
                sale__completed_at__range = (dt_from, dt_to),
            )
            .values("product", "product_name")
            .annotate(
                total_qty      = Sum("quantity"),
                total_revenue  = Sum("line_total"),
                total_tax      = Sum("sales_tax_applicable"),
                total_discount = Sum("discount_amount"),
                times_sold     = Count("sale", distinct=True),
            )
            .order_by("-total_revenue")[:self.limit]
        )
 
    def export_excel(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Sales"
        ws.freeze_panes = "A6"
 
        COLS     = 7
        next_row = _write_report_header(
            ws, "Product-wise Sales Report",
            self.company.business_name,
            self.date_from, self.date_to, COLS
        )
 
        headers = ["#", "Product Name", "Qty Sold", "Times in Sale",
                   "Revenue (Rs.)", "Tax (Rs.)", "Discount (Rs.)"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=next_row, column=i, value=h)
        _style_header_row(ws, next_row, COLS)
        next_row += 1
 
        data_start = next_row
        for i, row in enumerate(self.data, 1):
            r = next_row
            ws.cell(r, 1, i)
            ws.cell(r, 2, row["product_name"])
            ws.cell(r, 3, float(row["total_qty"]      or 0))
            ws.cell(r, 4, int(row["times_sold"]       or 0))
            ws.cell(r, 5, float(row["total_revenue"]  or 0))
            ws.cell(r, 6, float(row["total_tax"]      or 0))
            ws.cell(r, 7, float(row["total_discount"] or 0))
 
            ws.cell(r, 1).alignment = CENTER_ALIGN
            ws.cell(r, 2).alignment = LEFT_ALIGN
            for col in [3, 5, 6, 7]:
                ws.cell(r, col).number_format = PKR_FORMAT
            _style_data_row(ws, r, COLS, alternate=(i % 2 == 0))
            next_row += 1
 
        data_end = next_row - 1
 
        r = next_row
        ws.cell(r, 1, "TOTAL")
        ws.cell(r, 2, "")
        ws.cell(r, 3, f"=SUM(C{data_start}:C{data_end})")
        ws.cell(r, 4, f"=SUM(D{data_start}:D{data_end})")
        ws.cell(r, 5, f"=SUM(E{data_start}:E{data_end})")
        ws.cell(r, 6, f"=SUM(F{data_start}:F{data_end})")
        ws.cell(r, 7, f"=SUM(G{data_start}:G{data_end})")
        for col in [3, 5, 6, 7]:
            ws.cell(r, col).number_format = PKR_FORMAT
        _style_total_row(ws, r, COLS)
 
        _autofit_columns(ws)
        ws.sheet_view.showGridLines = False
 
        buffer = io.BytesIO()
        wb.save(buffer)
        return _save_to_s3(buffer, self.company, "product_sales", "xlsx")
 
    def export_pdf(self) -> str:
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=15*mm, leftMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm,
        )
        story = []
        width = landscape(A4)[0] - 30*mm
 
        story.append(Paragraph(
            self.company.business_name,
            ParagraphStyle("co", fontSize=12, fontName="Helvetica-Bold",
                           textColor=colors.HexColor("#1A56DB"),
                           alignment=TA_CENTER)
        ))
        story.append(Paragraph(
            "Product-wise Sales Report",
            ParagraphStyle("t", fontSize=14, fontName="Helvetica-Bold",
                           alignment=TA_CENTER, spaceAfter=4)
        ))
        story.append(Paragraph(
            f"Period: {self.date_from.strftime('%d %b %Y')} to "
            f"{self.date_to.strftime('%d %b %Y')}",
            ParagraphStyle("s", fontSize=9, alignment=TA_CENTER,
                           textColor=colors.HexColor("#6B7280"), spaceAfter=8)
        ))
 
        cs = ParagraphStyle("ch", fontSize=8, fontName="Helvetica-Bold",
                            textColor=colors.white, alignment=TA_CENTER)
        cr = ParagraphStyle("cr", fontSize=8, alignment=TA_RIGHT)
        cc = ParagraphStyle("cc", fontSize=8, alignment=TA_CENTER)
        cl = ParagraphStyle("cl", fontSize=8, alignment=TA_LEFT)
 
        table_data = [[
            Paragraph("#",              cs),
            Paragraph("Product Name",   cs),
            Paragraph("Qty Sold",       cs),
            Paragraph("Times in Sale",  cs),
            Paragraph("Revenue",        cs),
            Paragraph("Tax",            cs),
            Paragraph("Discount",       cs),
        ]]
 
        t_qty = t_rev = t_tax = t_dis = 0
        for i, row in enumerate(self.data, 1):
            qty = float(row["total_qty"]      or 0)
            rev = float(row["total_revenue"]  or 0)
            tax = float(row["total_tax"]      or 0)
            dis = float(row["total_discount"] or 0)
            t_qty += qty; t_rev += rev
            t_tax += tax; t_dis += dis
 
            table_data.append([
                Paragraph(str(i),                      cc),
                Paragraph(row["product_name"][:30],    cl),
                Paragraph(f"{qty:,.3f}",               cr),
                Paragraph(str(row["times_sold"] or 0), cc),
                Paragraph(f"Rs. {rev:,.2f}",           cr),
                Paragraph(f"Rs. {tax:,.2f}",           cr),
                Paragraph(f"Rs. {dis:,.2f}",           cr),
            ])
 
        tf = ParagraphStyle("tf", fontSize=8, fontName="Helvetica-Bold",
                            alignment=TA_RIGHT)
        tc = ParagraphStyle("tc", fontSize=8, fontName="Helvetica-Bold",
                            alignment=TA_CENTER)
        table_data.append([
            Paragraph("", tc), Paragraph("TOTAL", tc),
            Paragraph(f"{t_qty:,.3f}", tf),
            Paragraph("", tc),
            Paragraph(f"Rs. {t_rev:,.2f}", tf),
            Paragraph(f"Rs. {t_tax:,.2f}", tf),
            Paragraph(f"Rs. {t_dis:,.2f}", tf),
        ])
 
        col_w  = width / 7
        widths = [col_w*0.5, col_w*2.5, col_w, col_w, col_w, col_w, col_w]
        t = Table(table_data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0),  (-1, 0),  colors.HexColor("#1A56DB")),
            ("BACKGROUND",    (0, -1), (-1, -1), colors.HexColor("#DBEAFE")),
            ("ROWBACKGROUNDS",(0, 1),  (-1, -2),
             [colors.white, colors.HexColor("#EFF6FF")]),
            ("GRID",          (0, 0),  (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",    (0, 0),  (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0),  (-1, -1), 3),
            ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            f"Generated: {timezone.now().strftime('%d-%b-%Y %I:%M %p')}",
            ParagraphStyle("ft", fontSize=7, alignment=TA_RIGHT,
                           textColor=colors.HexColor("#9CA3AF"))
        ))
        doc.build(story)
        return _save_to_s3(buffer, self.company, "product_sales", "pdf")
 
 
# ===========================================================================
# 4. CASHIER-WISE SALES EXPORTER
# ===========================================================================
 
class CashierSalesExporter:
 
    def __init__(self, company, date_from: date, date_to: date):
        self.company   = company
        self.date_from = date_from
        self.date_to   = date_to
        self.data      = self._fetch_data()
 
    def _fetch_data(self) -> list:
        from pos.models import Sale, SaleStatus
        from django.db.models import Sum, Count, Avg
 
        dt_from = timezone.make_aware(
            datetime.combine(self.date_from, datetime.min.time())
        )
        dt_to   = timezone.make_aware(
            datetime.combine(self.date_to, datetime.max.time())
        )
 
        return list(
            Sale.objects.filter(
                company            = self.company,
                status             = SaleStatus.COMPLETED,
                completed_at__range = (dt_from, dt_to),
            )
            .values("cashier__id", "cashier__email",
                    "cashier__first_name", "cashier__last_name",
                    "cashier__role")
            .annotate(
                total_sales    = Count("id"),
                total_revenue  = Sum("total_amount"),
                total_tax      = Sum("total_tax"),
                avg_sale       = Avg("total_amount"),
            )
            .order_by("-total_revenue")
        )
 
    def export_excel(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cashier Sales"
        ws.freeze_panes = "A6"
 
        COLS     = 7
        next_row = _write_report_header(
            ws, "Cashier-wise Sales Report",
            self.company.business_name,
            self.date_from, self.date_to, COLS
        )
 
        headers = ["#", "Cashier Name", "Email", "Role",
                   "No. of Sales", "Total Revenue (Rs.)", "Avg Sale (Rs.)"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=next_row, column=i, value=h)
        _style_header_row(ws, next_row, COLS)
        next_row += 1
 
        data_start = next_row
        for i, row in enumerate(self.data, 1):
            r    = next_row
            name = (
                f"{row['cashier__first_name']} {row['cashier__last_name']}".strip()
                or row["cashier__email"]
            )
            ws.cell(r, 1, i)
            ws.cell(r, 2, name)
            ws.cell(r, 3, row["cashier__email"])
            ws.cell(r, 4, row["cashier__role"].title())
            ws.cell(r, 5, row["total_sales"])
            ws.cell(r, 6, float(row["total_revenue"] or 0))
            ws.cell(r, 7, float(row["avg_sale"]      or 0))
 
            ws.cell(r, 1).alignment = CENTER_ALIGN
            ws.cell(r, 4).alignment = CENTER_ALIGN
            for col in [6, 7]:
                ws.cell(r, col).number_format = PKR_FORMAT
            ws.cell(r, 5).number_format = INT_FORMAT
            _style_data_row(ws, r, COLS, alternate=(i % 2 == 0))
            next_row += 1
 
        data_end = next_row - 1
        r = next_row
        ws.cell(r, 1, "TOTAL")
        ws.cell(r, 5, f"=SUM(E{data_start}:E{data_end})")
        ws.cell(r, 6, f"=SUM(F{data_start}:F{data_end})")
        ws.cell(r, 7, f"=AVERAGE(G{data_start}:G{data_end})")
        for col in [6, 7]:
            ws.cell(r, col).number_format = PKR_FORMAT
        ws.cell(r, 5).number_format = INT_FORMAT
        _style_total_row(ws, r, COLS)
 
        _autofit_columns(ws)
        ws.sheet_view.showGridLines = False
 
        buffer = io.BytesIO()
        wb.save(buffer)
        return _save_to_s3(buffer, self.company, "cashier_sales", "xlsx")
 
    def export_pdf(self) -> str:
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=15*mm, leftMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm,
        )
        story = []
        width = A4[0] - 30*mm
 
        story.append(Paragraph(
            self.company.business_name,
            ParagraphStyle("co", fontSize=12, fontName="Helvetica-Bold",
                           textColor=colors.HexColor("#1A56DB"),
                           alignment=TA_CENTER)
        ))
        story.append(Paragraph(
            "Cashier-wise Sales Report",
            ParagraphStyle("t", fontSize=14, fontName="Helvetica-Bold",
                           alignment=TA_CENTER, spaceAfter=8)
        ))
 
        cs = ParagraphStyle("ch", fontSize=8, fontName="Helvetica-Bold",
                            textColor=colors.white, alignment=TA_CENTER)
        cr = ParagraphStyle("cr", fontSize=8, alignment=TA_RIGHT)
        cc = ParagraphStyle("cc", fontSize=8, alignment=TA_CENTER)
        cl = ParagraphStyle("cl", fontSize=8, alignment=TA_LEFT)
 
        table_data = [[
            Paragraph("#",           cs),
            Paragraph("Cashier",     cs),
            Paragraph("Role",        cs),
            Paragraph("No. Sales",   cs),
            Paragraph("Revenue",     cs),
            Paragraph("Avg Sale",    cs),
        ]]
 
        t_cnt = t_rev = 0
        for i, row in enumerate(self.data, 1):
            name = (
                f"{row['cashier__first_name']} {row['cashier__last_name']}".strip()
                or row["cashier__email"]
            )
            cnt = row["total_sales"]
            rev = float(row["total_revenue"] or 0)
            avg = float(row["avg_sale"]      or 0)
            t_cnt += cnt; t_rev += rev
 
            table_data.append([
                Paragraph(str(i), cc),
                Paragraph(name[:25], cl),
                Paragraph(row["cashier__role"].title(), cc),
                Paragraph(str(cnt), cc),
                Paragraph(f"Rs. {rev:,.2f}", cr),
                Paragraph(f"Rs. {avg:,.2f}", cr),
            ])
 
        avg_t = t_rev / t_cnt if t_cnt > 0 else 0
        tf = ParagraphStyle("tf", fontSize=8, fontName="Helvetica-Bold",
                            alignment=TA_RIGHT)
        tc = ParagraphStyle("tc", fontSize=8, fontName="Helvetica-Bold",
                            alignment=TA_CENTER)
        table_data.append([
            Paragraph("", tc), Paragraph("TOTAL", tc), Paragraph("", tc),
            Paragraph(str(t_cnt), tc),
            Paragraph(f"Rs. {t_rev:,.2f}", tf),
            Paragraph(f"Rs. {avg_t:,.2f}", tf),
        ])
 
        col_w  = width / 6
        widths = [col_w*0.4, col_w*2.2, col_w*0.8, col_w*0.7, col_w, col_w]
        t = Table(table_data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0),  (-1, 0),  colors.HexColor("#1A56DB")),
            ("BACKGROUND",    (0, -1), (-1, -1), colors.HexColor("#DBEAFE")),
            ("ROWBACKGROUNDS",(0, 1),  (-1, -2),
             [colors.white, colors.HexColor("#EFF6FF")]),
            ("GRID",          (0, 0),  (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",    (0, 0),  (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
            ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            f"Generated: {timezone.now().strftime('%d-%b-%Y %I:%M %p')}",
            ParagraphStyle("ft", fontSize=7, alignment=TA_RIGHT,
                           textColor=colors.HexColor("#9CA3AF"))
        ))
        doc.build(story)
        return _save_to_s3(buffer, self.company, "cashier_sales", "pdf")
 
 
# ===========================================================================
# 5. FBR SUBMISSION STATUS EXPORTER
# ===========================================================================
 
class FBRStatusExporter:
 
    def __init__(self, company, date_from: date, date_to: date):
        self.company   = company
        self.date_from = date_from
        self.date_to   = date_to
        self.data      = self._fetch_data()
 
    def _fetch_data(self) -> list:
        from pos.models import Sale, SaleStatus
 
        dt_from = timezone.make_aware(
            datetime.combine(self.date_from, datetime.min.time())
        )
        dt_to   = timezone.make_aware(
            datetime.combine(self.date_to, datetime.max.time())
        )
 
        return list(
            Sale.objects.filter(
                company            = self.company,
                status             = SaleStatus.COMPLETED,
                completed_at__range = (dt_from, dt_to),
            )
            .select_related("customer", "cashier")
            .order_by("-completed_at")
        )
 
    def export_excel(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "FBR Status"
        ws.freeze_panes = "A6"
 
        COLS     = 8
        next_row = _write_report_header(
            ws, "FBR Submission Status Report",
            self.company.business_name,
            self.date_from, self.date_to, COLS
        )
 
        headers = ["Sale #", "Date", "Customer", "Amount (Rs.)",
                   "FBR Status", "FBR Invoice #", "Scenario", "Error"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=next_row, column=i, value=h)
        _style_header_row(ws, next_row, COLS)
        next_row += 1
 
        STATUS_COLORS = {
            "success":  "D1FAE5",
            "pending":  "FEF3C7",
            "failed":   "FEE2E2",
            "skipped":  "F3F4F6",
        }
 
        for i, sale in enumerate(self.data):
            r = next_row
            ws.cell(r, 1, sale.sale_number)
            ws.cell(r, 2, sale.completed_at.strftime("%d-%b-%Y %H:%M")
                    if sale.completed_at else "")
            ws.cell(r, 3, sale.customer.name[:25])
            ws.cell(r, 4, float(sale.total_amount))
            ws.cell(r, 5, sale.fbr_submission_status.upper())
            ws.cell(r, 6, sale.fbr_invoice_number or "—")
            ws.cell(r, 7, sale.fbr_scenario_id    or "—")
            ws.cell(r, 8, sale.fbr_error_message[:50] if sale.fbr_error_message else "—")
 
            ws.cell(r, 4).number_format = PKR_FORMAT
            ws.cell(r, 5).alignment     = CENTER_ALIGN
 
            # Color by FBR status
            status_color = STATUS_COLORS.get(sale.fbr_submission_status, "FFFFFF")
            ws.cell(r, 5).fill = PatternFill("solid", fgColor=status_color)
            ws.cell(r, 5).font = Font(name="Arial", size=9, bold=True)
 
            _style_data_row(ws, r, COLS, alternate=False)
            next_row += 1
 
        _autofit_columns(ws)
        ws.sheet_view.showGridLines = False
 
        # Summary sheet
        ws2         = wb.create_sheet("Summary")
        from pos.models import FBRSubmissionStatus
        from django.db.models import Count, Sum
        from pos.models import Sale, SaleStatus
 
        dt_from = timezone.make_aware(
            datetime.combine(self.date_from, datetime.min.time())
        )
        dt_to   = timezone.make_aware(
            datetime.combine(self.date_to, datetime.max.time())
        )
 
        summary = list(
            Sale.objects.filter(
                company            = self.company,
                status             = SaleStatus.COMPLETED,
                completed_at__range = (dt_from, dt_to),
            )
            .values("fbr_submission_status")
            .annotate(count=Count("id"), total=Sum("total_amount"))
        )
 
        ws2["A1"] = "FBR Submission Summary"
        ws2["A1"].font = TITLE_FONT
        ws2.merge_cells("A1:C1")
 
        ws2["A3"] = "Status"
        ws2["B3"] = "Count"
        ws2["C3"] = "Total Amount (Rs.)"
        _style_header_row(ws2, 3, 3)
 
        for i, row in enumerate(summary, 4):
            ws2.cell(i, 1, row["fbr_submission_status"].upper())
            ws2.cell(i, 2, row["count"])
            ws2.cell(i, 3, float(row["total"] or 0))
            ws2.cell(i, 3).number_format = PKR_FORMAT
            _style_data_row(ws2, i, 3, alternate=(i % 2 == 0))
 
        _autofit_columns(ws2)
 
        buffer = io.BytesIO()
        wb.save(buffer)
        return _save_to_s3(buffer, self.company, "fbr_status", "xlsx")
 
    def export_pdf(self) -> str:
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=15*mm, leftMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm,
        )
        story = []
        width = landscape(A4)[0] - 30*mm
 
        story.append(Paragraph(
            self.company.business_name,
            ParagraphStyle("co", fontSize=12, fontName="Helvetica-Bold",
                           textColor=colors.HexColor("#1A56DB"),
                           alignment=TA_CENTER)
        ))
        story.append(Paragraph(
            "FBR Submission Status Report",
            ParagraphStyle("t", fontSize=14, fontName="Helvetica-Bold",
                           alignment=TA_CENTER, spaceAfter=8)
        ))
 
        cs = ParagraphStyle("ch", fontSize=7, fontName="Helvetica-Bold",
                            textColor=colors.white, alignment=TA_CENTER)
        cr = ParagraphStyle("cr", fontSize=7, alignment=TA_RIGHT)
        cc = ParagraphStyle("cc", fontSize=7, alignment=TA_CENTER)
        cl = ParagraphStyle("cl", fontSize=7, alignment=TA_LEFT)
 
        table_data = [[
            Paragraph("Sale #",      cs),
            Paragraph("Date",        cs),
            Paragraph("Customer",    cs),
            Paragraph("Amount",      cs),
            Paragraph("FBR Status",  cs),
            Paragraph("FBR Inv #",   cs),
            Paragraph("Scenario",    cs),
        ]]
 
        STATUS_COLORS_PDF = {
            "success": colors.HexColor("#D1FAE5"),
            "pending": colors.HexColor("#FEF3C7"),
            "failed":  colors.HexColor("#FEE2E2"),
            "skipped": colors.HexColor("#F3F4F6"),
        }
 
        table_styles = [
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1A56DB")),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]
 
        for i, sale in enumerate(self.data, 1):
            status_color = STATUS_COLORS_PDF.get(
                sale.fbr_submission_status, colors.white
            )
            table_styles.append((
                "BACKGROUND", (4, i), (4, i), status_color
            ))
            table_data.append([
                Paragraph(sale.sale_number,      cc),
                Paragraph(sale.completed_at.strftime("%d-%b-%Y")
                          if sale.completed_at else "", cc),
                Paragraph(sale.customer.name[:18], cl),
                Paragraph(f"Rs. {sale.total_amount:,.2f}", cr),
                Paragraph(sale.fbr_submission_status.upper(), cc),
                Paragraph((sale.fbr_invoice_number or "—")[:20], cc),
                Paragraph(sale.fbr_scenario_id or "—", cc),
            ])
 
        col_w  = width / 7
        widths = [col_w*1.2, col_w*0.9, col_w*1.3, col_w,
                  col_w*0.8, col_w*1.8, col_w]
        t = Table(table_data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle(table_styles))
        story.append(t)
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            f"Generated: {timezone.now().strftime('%d-%b-%Y %I:%M %p')}",
            ParagraphStyle("ft", fontSize=7, alignment=TA_RIGHT,
                           textColor=colors.HexColor("#9CA3AF"))
        ))
        doc.build(story)
        return _save_to_s3(buffer, self.company, "fbr_status", "pdf")
